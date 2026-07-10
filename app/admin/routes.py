import csv
import io
import secrets
from datetime import datetime, timedelta

import openpyxl
from flask import Blueprint, render_template, request, redirect, url_for, flash, Response, abort
from flask_login import login_required, current_user
from sqlalchemy import func, or_

from app.extensions import db
from app.models import (
    User, Profile, EmbaDetails, Batch, Company, Payment, ActivityLog, Announcement,
    ROLE_MEMBER, ROLE_ADMIN, ROLE_SUPER_ADMIN,
    STATUS_PENDING, STATUS_ACTIVE, STATUS_INACTIVE, STATUS_BLOCKED, STATUS_REJECTED, STATUS_UNREGISTERED,
    PAYMENT_SUCCESS, PAYMENT_TYPES, PAYMENT_METHODS,
)
from app.utils import admin_required, super_admin_required, log_admin_action, send_dev_email, new_token, is_strong_password

admin_bp = Blueprint("admin", __name__)


def _payment_status_for(user):
    today = datetime.utcnow().date()
    latest = (
        Payment.query.filter_by(user_id=user.id, status=PAYMENT_SUCCESS)
        .order_by(Payment.valid_until.desc())
        .first()
    )
    if not latest:
        return "Due"
    if latest.valid_until and latest.valid_until >= today:
        return "Paid"
    return "Due"


@admin_bp.route("/dashboard")
@login_required
@admin_required
def dashboard():
    total_members = User.query.filter(User.role == ROLE_MEMBER, User.status != STATUS_UNREGISTERED).count()
    pending = User.query.filter_by(role=ROLE_MEMBER, status=STATUS_PENDING).count()
    active = User.query.filter_by(role=ROLE_MEMBER, status=STATUS_ACTIVE).count()
    preloaded_count = User.query.filter_by(role=ROLE_MEMBER, status=STATUS_UNREGISTERED).count()

    now = datetime.utcnow()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    year_start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)

    month_total = db.session.query(func.coalesce(func.sum(Payment.amount), 0)).filter(
        Payment.status == PAYMENT_SUCCESS, Payment.paid_at >= month_start
    ).scalar()
    year_total = db.session.query(func.coalesce(func.sum(Payment.amount), 0)).filter(
        Payment.status == PAYMENT_SUCCESS, Payment.paid_at >= year_start
    ).scalar()

    batch_counts = (
        db.session.query(Batch.batch_name, func.count(EmbaDetails.user_id))
        .join(EmbaDetails, EmbaDetails.batch_id == Batch.id)
        .group_by(Batch.batch_name)
        .order_by(Batch.batch_name)
        .all()
    )

    return render_template(
        "admin/dashboard.html",
        total_members=total_members, pending=pending, active=active, preloaded_count=preloaded_count,
        month_total=month_total, year_total=year_total,
        batch_labels=[b[0] for b in batch_counts],
        batch_values=[b[1] for b in batch_counts],
    )


@admin_bp.route("/users")
@login_required
@admin_required
def users_list():
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "")
    batch = request.args.get("batch", "")

    query = User.query.filter_by(role=ROLE_MEMBER)
    if q:
        like = f"%{q}%"
        query = query.filter(or_(User.name.ilike(like), User.email.ilike(like)))
    if status:
        query = query.filter(User.status == status)
    else:
        # preloaded roster rows aren't real accounts yet — keep them out of the default view
        query = query.filter(User.status != STATUS_UNREGISTERED)
    if batch:
        query = query.join(EmbaDetails, EmbaDetails.user_id == User.id).join(Batch, EmbaDetails.batch_id == Batch.id).filter(Batch.batch_name == batch)

    members = query.order_by(User.created_at.desc()).all()
    batches = Batch.query.order_by(Batch.batch_name).all()

    rows = [(m, _payment_status_for(m)) for m in members]
    return render_template(
        "admin/users_list.html", rows=rows, batches=batches, q=q, status=status, batch=batch,
        statuses=[STATUS_UNREGISTERED, STATUS_PENDING, STATUS_ACTIVE, STATUS_INACTIVE, STATUS_BLOCKED, STATUS_REJECTED],
    )


@admin_bp.route("/users/<int:user_id>/approve", methods=["GET", "POST"])
@login_required
@admin_required
def approve_user(user_id):
    member = db.session.get(User, user_id)
    if not member:
        abort(404)
    batches = Batch.query.order_by(Batch.batch_name).all()

    if request.method == "POST":
        batch_id = request.form.get("batch_id")
        passing_year = request.form.get("passing_year")
        class_roll = request.form.get("class_roll")

        if not member.emba_details:
            member.emba_details = EmbaDetails(user_id=member.id)
        member.emba_details.batch_id = int(batch_id) if batch_id else None
        member.emba_details.passing_year = int(passing_year) if passing_year else None
        member.emba_details.class_roll = class_roll or None
        member.status = STATUS_ACTIVE
        db.session.commit()

        log_admin_action("approve_user", target_user_id=member.id)
        send_dev_email(
            to_email=member.email,
            subject="Your EMBA Member account has been approved",
            body_html=f"Hi {member.name}, your account has been approved. Please log in and complete payment to unlock directory access.",
            link_url=url_for("auth.login", _external=True),
        )
        flash(f"{member.name} approved.", "success")
        return redirect(url_for("admin.users_list"))

    return render_template("admin/approve_user.html", member=member, batches=batches)


@admin_bp.route("/users/<int:user_id>/reject", methods=["POST"])
@login_required
@admin_required
def reject_user(user_id):
    member = db.session.get(User, user_id)
    if not member:
        abort(404)
    reason = request.form.get("reason", "").strip()
    member.status = STATUS_REJECTED
    member.rejection_reason = reason
    db.session.commit()
    log_admin_action("reject_user", target_user_id=member.id, details=reason)
    send_dev_email(
        to_email=member.email,
        subject="Your EMBA Member registration was not approved",
        body_html=f"Hi {member.name}, unfortunately your registration was not approved. Reason: {reason or 'not specified'}.",
    )
    flash(f"{member.name} rejected.", "success")
    return redirect(url_for("admin.users_list"))


@admin_bp.route("/users/<int:user_id>/set-status", methods=["POST"])
@login_required
@admin_required
def set_status(user_id):
    member = db.session.get(User, user_id)
    if not member:
        abort(404)
    new_status = request.form.get("status")
    if new_status in (STATUS_ACTIVE, STATUS_INACTIVE, STATUS_BLOCKED):
        member.status = new_status
        db.session.commit()
        log_admin_action("set_status", target_user_id=member.id, details=new_status)
        flash(f"{member.name} is now {new_status}.", "success")
    return redirect(url_for("admin.users_list"))


@admin_bp.route("/users/<int:user_id>/reset-password", methods=["POST"])
@login_required
@admin_required
def reset_password(user_id):
    member = db.session.get(User, user_id)
    if not member:
        abort(404)
    new_pw = secrets.token_urlsafe(9) + "!A1"
    member.set_password(new_pw)
    db.session.commit()
    log_admin_action("reset_password", target_user_id=member.id)
    send_dev_email(
        to_email=member.email,
        subject="Your EMBA Member password has been reset",
        body_html=f"Hi {member.name}, an admin reset your password. Your temporary password is: {new_pw} — please change it after logging in.",
        link_url=url_for("auth.login", _external=True),
    )
    flash(f"Password reset for {member.name}; new credentials emailed (see Dev Inbox).", "success")
    return redirect(url_for("admin.users_list"))


@admin_bp.route("/users/<int:user_id>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def edit_user(user_id):
    member = db.session.get(User, user_id)
    if not member:
        abort(404)
    if not member.profile:
        member.profile = Profile(user_id=member.id)
    if not member.emba_details:
        member.emba_details = EmbaDetails(user_id=member.id)
    db.session.commit()

    batches = Batch.query.order_by(Batch.batch_name).all()

    if request.method == "POST":
        member.name = request.form.get("name", member.name).strip()
        member.profile.mobile_personal = request.form.get("mobile_personal") or None
        member.profile.current_designation = request.form.get("current_designation") or None
        member.profile.location = request.form.get("location") or None
        batch_id = request.form.get("batch_id")
        member.emba_details.batch_id = int(batch_id) if batch_id else None
        member.emba_details.passing_year = int(request.form["passing_year"]) if request.form.get("passing_year") else None
        member.emba_details.class_roll = request.form.get("class_roll") or None
        db.session.commit()
        log_admin_action("edit_profile", target_user_id=member.id)
        flash("Profile updated.", "success")
        return redirect(url_for("admin.users_list"))

    return render_template("admin/edit_user.html", member=member, batches=batches)


@admin_bp.route("/users/new", methods=["GET", "POST"])
@login_required
@admin_required
def new_user():
    batches = Batch.query.order_by(Batch.batch_name).all()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip().lower()
        mobile = request.form.get("mobile", "").strip()
        batch_id = request.form.get("batch_id")
        passing_year = request.form.get("passing_year")
        class_roll = request.form.get("class_roll")

        if User.query.filter_by(email=email).first():
            flash("A user with this email already exists.", "error")
            return render_template("admin/new_user.html", batches=batches)

        temp_pw = secrets.token_urlsafe(9) + "!A1"
        user = User(name=name, email=email, role=ROLE_MEMBER, status=STATUS_ACTIVE, email_verified_at=datetime.utcnow())
        user.set_password(temp_pw)
        db.session.add(user)
        db.session.flush()
        user.profile = Profile(user_id=user.id, mobile_personal=mobile or None)
        user.emba_details = EmbaDetails(
            user_id=user.id,
            batch_id=int(batch_id) if batch_id else None,
            passing_year=int(passing_year) if passing_year else None,
            class_roll=class_roll or None,
        )
        db.session.commit()
        log_admin_action("add_member_manually", target_user_id=user.id)
        send_dev_email(
            to_email=user.email,
            subject="Your EMBA Member account has been created",
            body_html=f"Hi {name}, an account was created for you. Login email: {email}, temporary password: {temp_pw}.",
            link_url=url_for("auth.login", _external=True),
        )
        flash(f"{name} added; login credentials emailed (see Dev Inbox).", "success")
        return redirect(url_for("admin.users_list"))

    return render_template("admin/new_user.html", batches=batches)


@admin_bp.route("/users/import", methods=["GET", "POST"])
@login_required
@admin_required
def import_users():
    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename:
            flash("Please choose a file.", "error")
            return redirect(url_for("admin.import_users"))

        rows = []
        if file.filename.lower().endswith(".csv"):
            content = file.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
        elif file.filename.lower().endswith((".xlsx", ".xlsm")):
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        else:
            flash("Unsupported file type. Please upload .csv or .xlsx.", "error")
            return redirect(url_for("admin.import_users"))

        created, skipped = 0, 0
        for row in rows:
            email = str(row.get("email") or "").strip().lower()
            name = str(row.get("name") or "").strip()
            if not email or not name or User.query.filter_by(email=email).first():
                skipped += 1
                continue
            temp_pw = secrets.token_urlsafe(9) + "!A1"
            user = User(name=name, email=email, role=ROLE_MEMBER, status=STATUS_ACTIVE, email_verified_at=datetime.utcnow())
            user.set_password(temp_pw)
            db.session.add(user)
            db.session.flush()
            batch_name = str(row.get("batch") or "").strip()
            batch = Batch.query.filter_by(batch_name=batch_name).first() if batch_name else None
            user.profile = Profile(user_id=user.id, mobile_personal=str(row.get("mobile") or "").strip() or None)
            user.emba_details = EmbaDetails(user_id=user.id, batch_id=batch.id if batch else None, class_roll=str(row.get("roll") or "").strip() or None)
            send_dev_email(
                to_email=user.email,
                subject="Your EMBA Member account has been created (bulk import)",
                body_html=f"Hi {name}, an account was created for you. Login email: {email}, temporary password: {temp_pw}.",
                link_url=url_for("auth.login", _external=True),
            )
            created += 1
        db.session.commit()
        log_admin_action("bulk_import", details=f"created={created}, skipped={skipped}")
        flash(f"Import complete: {created} members created, {skipped} skipped (duplicate/invalid).", "success")
        return redirect(url_for("admin.users_list"))

    return render_template("admin/import_users.html")


@admin_bp.route("/users/export")
@login_required
@admin_required
def export_users():
    log_admin_action("export_users")
    members = User.query.filter_by(role=ROLE_MEMBER).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Name", "Email", "Batch", "Passing Year", "Roll", "Mobile", "Status", "Payment Status"])
    for m in members:
        writer.writerow([
            m.name, m.email,
            m.emba_details.batch.batch_name if m.emba_details and m.emba_details.batch else "",
            m.emba_details.passing_year if m.emba_details else "",
            m.emba_details.class_roll if m.emba_details else "",
            m.profile.mobile_personal if m.profile else "",
            m.status, _payment_status_for(m),
        ])
    return Response(output.getvalue(), mimetype="text/csv",
                     headers={"Content-Disposition": "attachment; filename=members_export.csv"})


@admin_bp.route("/payments", methods=["GET", "POST"])
@login_required
@admin_required
def payments():
    if request.method == "POST":
        user_id = int(request.form["user_id"])
        amount = float(request.form["amount"])
        payment_type = request.form["payment_type"]
        method = request.form["method"]
        reference_no = request.form.get("reference_no", "").strip()
        note = request.form.get("note", "").strip()
        paid_at = datetime.utcnow()

        payment = Payment(
            user_id=user_id, amount=amount, payment_type=payment_type, method=method,
            reference_no=reference_no or None, note=note or None, status=PAYMENT_SUCCESS,
            paid_at=paid_at, valid_until=(paid_at + timedelta(days=365)).date(),
            recorded_by=current_user.id,
        )
        db.session.add(payment)
        db.session.commit()
        log_admin_action("record_payment", target_user_id=user_id, details=f"{amount} via {method}")
        flash("Payment recorded.", "success")
        return redirect(url_for("admin.payments"))

    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    batch = request.args.get("batch")
    payment_type = request.args.get("payment_type")

    query = Payment.query.join(User, Payment.user_id == User.id)
    if date_from:
        query = query.filter(Payment.created_at >= date_from)
    if date_to:
        query = query.filter(Payment.created_at <= date_to + " 23:59:59")
    if batch:
        query = query.join(EmbaDetails, EmbaDetails.user_id == User.id).join(Batch, EmbaDetails.batch_id == Batch.id).filter(Batch.batch_name == batch)
    if payment_type:
        query = query.filter(Payment.payment_type == payment_type)

    payment_rows = query.order_by(Payment.created_at.desc()).all()
    total = sum(p.amount for p in payment_rows if p.status == PAYMENT_SUCCESS)

    members = User.query.filter_by(role=ROLE_MEMBER).order_by(User.name).all()
    batches = Batch.query.order_by(Batch.batch_name).all()

    return render_template(
        "admin/payments.html", payment_rows=payment_rows, total=total, members=members, batches=batches,
        payment_types=PAYMENT_TYPES, payment_methods=PAYMENT_METHODS,
        date_from=date_from, date_to=date_to, batch=batch, payment_type=payment_type,
    )


@admin_bp.route("/payments/export")
@login_required
@admin_required
def export_payments():
    log_admin_action("export_payments")
    rows = Payment.query.order_by(Payment.created_at.desc()).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Member", "Email", "Amount", "Type", "Method", "Status", "Reference", "Valid Until"])
    for p in rows:
        writer.writerow([p.created_at.strftime("%Y-%m-%d"), p.user.name, p.user.email, p.amount, p.payment_type,
                          p.method, p.status, p.reference_no or "", p.valid_until or ""])
    return Response(output.getvalue(), mimetype="text/csv",
                     headers={"Content-Disposition": "attachment; filename=payments_export.csv"})


@admin_bp.route("/batches", methods=["GET", "POST"])
@login_required
@admin_required
def batches():
    if request.method == "POST":
        batch_id = request.form.get("batch_id")
        batch_name = request.form.get("batch_name", "").strip()
        passing_year = request.form.get("passing_year")

        if batch_id:
            b = db.session.get(Batch, int(batch_id))
            b.batch_name = batch_name
            b.passing_year = int(passing_year) if passing_year else None
            b.passing_year_placeholder = False
        else:
            b = Batch(batch_name=batch_name, passing_year=int(passing_year) if passing_year else None)
            db.session.add(b)
        db.session.commit()
        log_admin_action("edit_batch", details=batch_name)
        flash("Batch saved.", "success")
        return redirect(url_for("admin.batches"))

    all_batches = Batch.query.order_by(Batch.batch_name).all()
    return render_template("admin/batches.html", batches=all_batches)


@admin_bp.route("/announcements", methods=["GET", "POST"])
@login_required
@admin_required
def announcements():
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        body = request.form.get("body", "").strip()
        if title and body:
            db.session.add(Announcement(title=title, body=body, created_by=current_user.id))
            db.session.commit()
            flash("Announcement posted.", "success")
        return redirect(url_for("admin.announcements"))

    all_announcements = Announcement.query.order_by(Announcement.created_at.desc()).all()
    return render_template("admin/announcements.html", announcements=all_announcements)


@admin_bp.route("/announcements/<int:ann_id>/toggle", methods=["POST"])
@login_required
@admin_required
def toggle_announcement(ann_id):
    ann = db.session.get(Announcement, ann_id)
    if ann:
        ann.active = not ann.active
        db.session.commit()
    return redirect(url_for("admin.announcements"))


@admin_bp.route("/admins", methods=["GET", "POST"])
@login_required
@super_admin_required
def admins():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip().lower()
        if User.query.filter_by(email=email).first():
            flash("A user with this email already exists.", "error")
        else:
            temp_pw = secrets.token_urlsafe(9) + "!A1"
            admin_user = User(name=name, email=email, role=ROLE_ADMIN, status=STATUS_ACTIVE, email_verified_at=datetime.utcnow())
            admin_user.set_password(temp_pw)
            db.session.add(admin_user)
            db.session.commit()
            log_admin_action("create_admin", target_user_id=admin_user.id)
            send_dev_email(
                to_email=email,
                subject="You've been made an EMBA Member admin",
                body_html=f"Hi {name}, you now have admin access. Login email: {email}, temporary password: {temp_pw}.",
                link_url=url_for("auth.login", _external=True),
            )
            flash(f"{name} added as admin; credentials emailed (see Dev Inbox).", "success")
        return redirect(url_for("admin.admins"))

    admin_users = User.query.filter(User.role.in_([ROLE_ADMIN, ROLE_SUPER_ADMIN])).order_by(User.created_at).all()
    return render_template("admin/admins.html", admin_users=admin_users)


@admin_bp.route("/admins/<int:user_id>/remove", methods=["POST"])
@login_required
@super_admin_required
def remove_admin(user_id):
    target = db.session.get(User, user_id)
    if target and target.role == ROLE_ADMIN and target.id != current_user.id:
        target.status = STATUS_BLOCKED
        db.session.commit()
        log_admin_action("remove_admin", target_user_id=target.id)
        flash(f"{target.name} removed as admin.", "success")
    return redirect(url_for("admin.admins"))


@admin_bp.route("/activity-log")
@login_required
@admin_required
def activity_log():
    logs = ActivityLog.query.order_by(ActivityLog.created_at.desc()).limit(300).all()
    return render_template("admin/activity_log.html", logs=logs)
